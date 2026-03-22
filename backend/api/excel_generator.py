"""
Excel File Generator

Creates formatted Excel files from paper analysis results.
Includes:
- All extracted fields in columns
- Header formatting
- Column width adjustment
- Multiple papers in multiple rows
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
from datetime import datetime


def generate_excel(results: List[Dict[str, Any]], spanish_results: List[Dict[str, Any]], output_path: str) -> str:
    """
    Generate Excel file from paper analysis results.
    
    Args:
        results: List of dictionaries with English analysis data for each paper
        spanish_results: List of dictionaries with Spanish analysis data for each paper
        output_path: Where to save the Excel file
        
    Returns:
        Path to the generated Excel file
        
    The Excel file will have two sheets:
    Sheet 1 – "Research Papers Analysis": all columns in English
    Sheet 2 – "Análisis en Español": same columns and data in Spanish
    """
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Research Papers Analysis"
    
    # Define columns (in order)
    columns = [
        ("file_name", "File Name"),
        ("authors", "Authors"),
        ("publication_year", "Publication Year"),
        ("journal", "Journal"),
        ("doi", "DOI"),
        ("country_of_publication", "Country of Publication"),
        ("title", "Title"),
        ("paper_type", "Type of Paper"),
        ("objective", "Objective"),
        ("methodology_summary", "Methodology Summary"),
        ("study_sample_size", "Study Sample Size"),
        ("study_duration", "Study Duration"),
        ("inclusion_criteria", "Inclusion Criteria"),
        ("country_of_population", "Country of Population"),
        ("conclusion", "Conclusion"),
    ]
    
    # Header styling
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_idx, (_, header_name) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    for row_idx, result in enumerate(results, start=2):
        for col_idx, (field_key, _) in enumerate(columns, start=1):
            # Get value from result
            value = result.get(field_key, "")
            
            # Format authors as comma-separated string
            if field_key == "authors" and isinstance(value, list):
                value = ", ".join(value)
            
            # Handle None values
            if value is None:
                value = "N/A"
            
            # Write cell
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border
    
    # Adjust column widths
    column_widths = {
        "A": 20,  # File Name
        "B": 40,  # Authors
        "C": 15,  # Year
        "D": 30,  # Journal
        "E": 25,  # DOI
        "F": 20,  # Country of Publication
        "G": 50,  # Title
        "H": 15,  # Type
        "I": 60,  # Objective
        "J": 60,  # Methodology
        "K": 15,  # Sample Size
        "L": 15,  # Duration
        "M": 40,  # Inclusion Criteria
        "N": 20,  # Country of Population
        "O": 60,  # Conclusion
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add Spanish sheet (same data, headers in Spanish)
    spanish_columns = [
        ("file_name", "Nombre de Archivo"),
        ("authors", "Autores"),
        ("publication_year", "Año de Publicación"),
        ("journal", "Revista"),
        ("doi", "DOI"),
        ("country_of_publication", "País de Publicación"),
        ("title", "Título"),
        ("paper_type", "Tipo de Artículo"),
        ("objective", "Objetivo"),
        ("methodology_summary", "Resumen de Metodología"),
        ("study_sample_size", "Tamaño de Muestra"),
        ("study_duration", "Duración del Estudio"),
        ("inclusion_criteria", "Criterios de Inclusión"),
        ("country_of_population", "País de la Población"),
        ("conclusion", "Conclusión"),
    ]

    es_ws = wb.create_sheet("Análisis en Español")

    # Write Spanish headers
    for col_idx, (_, header_name) in enumerate(spanish_columns, start=1):
        cell = es_ws.cell(row=1, column=col_idx, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data rows (Spanish translated data)
    for row_idx, result in enumerate(spanish_results, start=2):
        for col_idx, (field_key, _) in enumerate(spanish_columns, start=1):
            value = result.get(field_key, "")
            if field_key == "authors" and isinstance(value, list):
                value = ", ".join(value)
            if value is None:
                value = "N/A"
            cell = es_ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border

    # Apply same column widths
    for col_letter, width in column_widths.items():
        es_ws.column_dimensions[col_letter].width = width

    # Freeze header row
    es_ws.freeze_panes = 'A2'
    
    # Save
    wb.save(output_path)
    
    return output_path


def format_cell_value(value: Any) -> str:
    """
    Format a cell value for Excel display.
    
    Handles:
    - Lists (join with commas)
    - None (convert to "N/A")
    - Numbers (keep as is)
    - Strings (keep as is)
    """
    if value is None:
        return "N/A"
    elif isinstance(value, list):
        return ", ".join(str(item) for item in value)
    elif isinstance(value, (int, float)):
        return value
    else:
        return str(value)
